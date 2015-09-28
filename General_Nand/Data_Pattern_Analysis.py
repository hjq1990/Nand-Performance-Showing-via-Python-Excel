__author__ = '20093'
'''
--IPython Code for Data Pattern Frequency Analysis------------

--By Jinqiang He,   Sept 27th, 2015

'''

from functools import partial
import os
from openpyxl import Workbook
import wx

def get_path(wildcard):
    app = wx.App(None)
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    dialog = wx.FileDialog(None, 'Open', wildcard=wildcard, style=style)
    if dialog.ShowModal() == wx.ID_OK:
        path = dialog.GetPath()
    else:
        path = None
    dialog.Destroy()
    return path

def init():
    global ws0_row, ws0_col
    for i in xrange(256):
        ws0.cell(row=ws0_row+i,column=1).value=i
    for j in xrange(256):
        ws0.cell(row=1,column=ws0_col+j).value=j

def main():
    global excel_name,ws0_row, ws0_col
    print "Please select the file for the data pattern to be processed."
    filename = get_path('*.pat')
    print "Processing File", filename
    Freq=[[0 for x in xrange(300)] for y in xrange(300)]
    bytes_per_line=1
    byte_num=0
    excel_name = os.path.join(os.path.dirname(
        filename),os.path.basename(filename)[:-4]+ '--Analysis.xlsx')
    with open(filename, 'rb') as binary_file:
        for block in iter(partial(binary_file.read, bytes_per_line), ''):
            s1 = ' '.join('{0:02x}'.format(ord(b)) for b in block)

            if byte_num % 0x4650 <0x4490:
                Freq[int(s1,16)][byte_num / 0x4650]=Freq[int(s1,16)][byte_num / 0x4650]+1
            else:
                bb=0
            byte_num=byte_num+1
    for i in xrange(256):
        ws0.cell(row=ws0_col+i,column=2).value=hex(i)
        for j in xrange(256):
            ws0.cell(row=ws0_col+i,column=3+j).value=Freq[i][j]


if __name__ == '__main__':
    excel_name = ' '
    file = ' '
    ws0_row = 2
    ws0_col = 2
    print "This API is for auto analysis data pat files and result will be saved in an excel file of same basename."
    print "by Jinqiang Sept 27th, 2015"
    print '****************************************************'
    print ' '

#------------------Excel WorkBook Initiation----------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    ws0 = wb.create_sheet()
    ws0.title = "Freq-By-Page"
    init()
    main()
    print excel_name
    wb.save(excel_name)

    print "The processed Error Log Summary is saved as", excel_name