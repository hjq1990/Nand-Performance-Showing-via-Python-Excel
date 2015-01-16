'''
--IPython Code for Automatic Processing T Data of ORT eSSD-------------
--1: Iterate through all txt files and get data accordingly;
--2: Based on main functions {Erase(), FBC(),Program(),Read()...}, program process data and store into the same excel file;
--3: Max, Min, and Average data of Nand Performance are calculated and shown as a summary 


--By Jinqiang He,   August 17th, 2014

'''
from Tkinter import *
import tkSimpleDialog
from tkFileDialog import *
import glob
import os
from openpyxl import Workbook


def Erase(lines):
    global Erase_num,file,erase_time
    line=lines
    #print line        
    Erase_num=Erase_num+1

    ws0.cell(row =Erase_num +5, column = 1).value=file
    ws0.cell(row =Erase_num +5, column = 2).value=line[1]   
    ws0.cell(row =Erase_num +5, column = 3).value=float(line[6])
    if(len(line)>9):
        erase_time=1
        ws0.cell(row =Erase_num +5, column = 4).value=float(line[10])
        
def FBC(lines):
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    global FBC_event,file
    FBC_event=FBC_event+1
    ws3.cell(row = FBC_event+5, column = 1).value=file
    #ws3.cell(row = FBC_event+5, column = 2).value=line
    ws3.cell(row = FBC_event+5, column = 3).value=float(line[2])

def Program(lines):
    global PL,PU,file,err_num
    pages=int(lines[1] ,16)
    page= pages % 256
    line=lines
    if(line[7]!='0'):
        
        if((page==0) or ((page % 2 !=0) and (page!=255))):
            PL=PL+1
            ws1.cell(row = PL+5, column = 1).value=file
            ws1.cell(row = PL+5, column = 2).value=line[1]
            ws1.cell(row = PL+5, column = 3).value=float(line[7])
            ws1.cell(row = PL+5, column = 4).value=float(line[9])
                
        elif((page==255) or ((page % 2 ==0) and (page!=0))):
            PU=PU+1 
            ws1.cell(row = PU+5, column = 6).value=file
            ws1.cell(row = PU+5, column = 7).value=line[1]
            ws1.cell(row = PU+5, column = 8).value=float(line[7])
            ws1.cell(row = PU+5, column = 9).value=float(line[9])
        
def Read(lines):
    global Read_num,file
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    page=int(line[1][-2:],16)
    #print line
    global RL, RU
    if((page==0) or ((page % 2 !=0) and (page!=255))):
        RL=RL+1
        ws2.cell(row = RL+5, column = 1).value=file
        ws2.cell(row = RL+5, column = 2).value=line[1] 
        ws2.cell(row = RL+5, column = 3).value=float(line[4])     
        
    elif((page==255) or ((page % 2 ==0) and (page!=0))):
        RU=RU+1 
        ws2.cell(row = RU+5, column = 5).value=file
        ws2.cell(row = RU+5, column = 6).value=line[1]
        ws2.cell(row = RU+5, column = 7).value=float(line[4])

def init():
   
    ws0.cell('B2').value='Max'
    ws0.cell('B3').value='Min'
    ws0.cell('B4').value='Average'    
    ws0.cell('C1').value='Erase Loop'
    ws0.cell('D1').value='Erase Time'
    ws0.cell('A5').value='file'
    ws0.cell('B5').value='page'

    ws1.cell('B2').value='Max'
    ws1.cell('B3').value='Min'
    ws1.cell('B4').value='Average'
    ws1.cell('C1').value='P_L Loop'
    ws1.cell('D1').value='P_L Time'    
    ws1.cell('H1').value='P_U Loop'
    ws1.cell('I1').value='P_U Time' 
    ws1.cell('A5').value='file'
    ws1.cell('B5').value='page'    
    ws1.cell('F5').value='file'
    ws1.cell('G5').value='page' 
    
    ws2.cell('B2').value='Max'
    ws2.cell('B3').value='Min'
    ws2.cell('B4').value='Average'    
    ws2.cell('C1').value='RL Time'
    ws2.cell('G1').value='RU Time'
    ws2.cell('A5').value='file'
    ws2.cell('B6').value='page'
    ws2.cell('E5').value='file'
    ws2.cell('F5').value='page'
       
    ws3.cell('C1').value='FBC Check'
    ws3.cell('B2').value='FBC Max'
    ws3.cell('B3').value='FBC Min'
    ws3.cell('B4').value='FBC Average'
    ws3.cell('A5').value='file'
    #ws3.cell('B5').value='page'
       
    ws.cell('C3').value='Max'
    ws.cell('D3').value='Min'
    ws.cell('E3').value='Average'
    ws.cell('B4').value='Erase Loop'
    ws.cell('B5').value='Erase Time'  
    ws.cell('B6').value='PL Loop'
    ws.cell('B7').value='PL Time'
    ws.cell('B8').value='PU Loop'
    ws.cell('B9').value='PU Time'   
    ws.cell('B10').value='Read Low'
    ws.cell('B11').value='Read Up'  
    ws.cell('B12').value='FBC'
    
    ws.cell('B20').value='Error Event List'
    ws.cell('B21').value='err_num'
    ws.cell('C21').value='file'
    ws.cell('D21').value='line number'
    ws.cell('E21').value='err_item'

def Cal():
    global FBC_event,PL,PU,RL,RU,Erase_num
    
    ws0.cell('C2').value="=max(C6:C"+str(5+Erase_num)+')'
    ws0.cell('C3').value="=min(C6:C"+str(5+Erase_num)+')'
    ws0.cell('C4').value="=average(C6:C"+str(5+Erase_num)+')'
    
    if(erase_time==1):
        
        ws0.cell('D2').value="=max(D6:D"+str(5+Erase_num)+')'
        ws0.cell('D3').value="=min(D6:D"+str(5+Erase_num)+')'
        ws0.cell('D4').value="=average(D6:D"+str(5+Erase_num)+')'
        ws.cell('C5').value='=erase!$D$2'
        ws.cell('D5').value='=erase!$D$3'
        ws.cell('E5').value='=erase!$D$4'
    else:
        ws0.cell('D2').value="there is no erase time data for all data files"
        ws.cell('C5').value="there is no erase time data for all data files"
    

    ws1.cell('C2').value="=max(C6:C"+str(5+PL)+')'
    ws1.cell('C3').value="=min(C6:C"+str(5+PL)+')'
    ws1.cell('C4').value="=average(C6:C"+str(5+PL)+')'  
    ws1.cell('D2').value="=max(D6:D"+str(5+PL)+')'
    ws1.cell('D3').value="=min(D6:D"+str(5+PL)+')'
    ws1.cell('D4').value="=average(D6:D"+str(5+PL)+')'  
    
    ws1.cell('H2').value="=max(H6:H"+str(5+PU)+')'
    ws1.cell('H3').value="=min(H6:H"+str(5+PU)+')'
    ws1.cell('H4').value="=average(H6:H"+str(5+PU)+')'
    ws1.cell('I2').value="=max(I6:I"+str(5+PU)+')'
    ws1.cell('I3').value="=min(I6:I"+str(5+PU)+')'
    ws1.cell('I4').value="=average(I6:I"+str(5+PU)+')'              
    
    ws2.cell('C2').value="=max(C6:C"+str(5+RL)+')'
    ws2.cell('C3').value="=min(C6:C"+str(5+RL)+')'
    ws2.cell('C4').value="=average(C6:C"+str(5+RL)+')'
    ws2.cell('G2').value="=max(G6:G"+str(5+RU)+')'
    ws2.cell('G3').value="=min(G6:G"+str(5+RU)+')'
    ws2.cell('G4').value="=average(G6:G"+str(5+RU)+')'       
    
    ws3.cell('C2').value='=max'+"(C6:C"+str(5+FBC_event)+')'
    ws3.cell('C3').value='=min'+"(C6:C"+str(5+FBC_event)+')'
    ws3.cell('C4').value='=average'+"(C6:C"+str(5+FBC_event)+')'
    
    ws.cell('C4').value='=erase!$C$2'
    ws.cell('D4').value='=erase!$C$3'
    ws.cell('E4').value='=erase!$C$4'

    
    ws.cell('C6').value='=PROGRAM!$C$2'
    ws.cell('D6').value='=PROGRAM!$C$3'
    ws.cell('E6').value='=PROGRAM!$C$4'
    ws.cell('C7').value='=PROGRAM!$D$2'
    ws.cell('D7').value='=PROGRAM!$D$3'
    ws.cell('E7').value='=PROGRAM!$D$4'
    
    ws.cell('C8').value='=PROGRAM!$H$2'
    ws.cell('D8').value='=PROGRAM!$H$3'
    ws.cell('E8').value='=PROGRAM!$H$4'
    ws.cell('C9').value='=PROGRAM!$I$2'
    ws.cell('D9').value='=PROGRAM!$I$3'
    ws.cell('E9').value='=PROGRAM!$I$4'
    
    ws.cell('C10').value='=read!$c$2'
    ws.cell('D10').value='=read!$c$3'
    ws.cell('E10').value='=read!$c$4'
    
    ws.cell('C11').value='=read!$G$2'
    ws.cell('D11').value='=read!$G$3'
    ws.cell('E11').value='=read!$G$4'
    
    ws.cell('C12').value='=read!$C$2'
    ws.cell('D12').value='=read!$C$3'
    ws.cell('E12').value='=read!$C$4'
         
def Summarize():
    print 'final proces'

    
def Summarize2():
    aa;
    
 
      
   
             
def main(): 
    err_num=0
    global excel_name,file
    
    root = Tk()
    root.wm_title("T data")
    
    w = Label(root, text="Please select the folder for the T-data to be processed.") 
   
    dirname = askdirectory()
    print "Selecting folder:",dirname
    
    fold=dirname.replace('/',' ').split()
    excel_name=fold[len(fold)-1]+' T data.xlsx'
    print "Please close this dialog by clicking X on the right-top"
    
    w.pack()
    root.mainloop()
    os.chdir(dirname)
    for file in glob.glob("*.txt"):
        print "Processing File",file              
        f = open(file, 'r')    #open T.txx
        line_num=0
        for line in f:
            line_num=line_num+1
            if(line[32:44]=='Program Loop'):
                lines=line.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
                if(lines[4]=='E0'):
                    Program(lines)
                else:
                    #print line
                    err_num=err_num+1
                    ws.cell(row = 21+err_num, column = 2).value=err_num
                    ws.cell(row = 21+err_num, column = 3).value=file
                    ws.cell(row = 21+err_num, column = 4).value=line_num
                    ws.cell(row = 21+err_num, column = 5).value=line
                    
                continue
                
            if(line[5:10]=='Total'):
                FBC(line)
                continue
                
            if ((line[2:7]=='Block') and (len(line)>20)):             
                
                lines=line.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
                if(lines[4]=='E0'):
                    Erase(lines)
                else:
                    #print line
                    err_num=err_num+1
                    ws.cell(row = 21+err_num, column = 2).value=err_num
                    ws.cell(row = 21+err_num, column = 3).value=file
                    ws.cell(row = 21+err_num, column = 4).value=line_num
                    ws.cell(row = 21+err_num, column = 5).value=line
                continue
                
            if ((line[0:4]=='Page') and (len(line)<40)) :
                lines=line.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
                if(lines[4]=='E1'):
                    #print line
                    err_num=err_num+1
                    ws.cell(row = 21+err_num, column = 2).value=err_num
                    ws.cell(row = 21+err_num, column = 3).value=file
                    ws.cell(row = 21+err_num, column = 4).value=line_num
                    ws.cell(row = 21+err_num, column = 5).value=line
                else:
                    Read(line)
                continue
    if(err_num>0):
        print "kind suggestion:"
        print "Please further check some source data, some items may be wrong"
        print "For details, please refer to sheet 0:Summary in the final excel file"
        ws.cell('B16').value="kind suggestion:"
        ws.cell('B17').value="Please further check source data, some files may contain wrong data. You need to correctify failure items or remove the file from this folder" 
        ws.cell('B18').value="For details, please refer to below lines"    
    print '****************************************************'                         
    print "Congratulations!!!"
    print "Processing Finished, and data is saved as excel file as below:"
    print dirname+'\/'+excel_name         
if __name__ == '__main__':
    RL=0
    RU=0
    Erase_num=0
    FBC_event=0;
    Row_offset=4
    PL=0
    PU=0
    excel_name=' '
    file=' '
    erase_time=0


    
#------------------Excel WorkBook Initiation----------------------------------    
    wb=Workbook()
    ws = wb.active
    ws.title="Summary"
    ws0 = wb.create_sheet()
    ws0.title="erase"
    ws1 = wb.create_sheet()
    ws1.title="program"
    ws2 = wb.create_sheet()
    ws2.title="read"
    ws3 = wb.create_sheet()
    ws3.title="FBC"
    
    init()
    print "This API is for auto processing T data from txt files generated from NanoNT, and processed data will be saved in an excel file."
    
    print '****************************************************' 
    "Please select the folder for the T-data to be processed."
    main()
    Cal()
    #excel_name = tkSimpleDialog.askstring("Same Folder", "enter your name")
    #print excel_name
    #file_path=dirname+'\\'+str(excel_name)+'.xlsx'
    #print file_path
    wb.save(excel_name)
    #Summarize()
    
