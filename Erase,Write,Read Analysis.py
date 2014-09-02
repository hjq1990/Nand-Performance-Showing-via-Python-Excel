
from openpyxl import Workbook

def Erase(lines):
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    global Erase_num
   
    if(line[6]!='0'):
        Erase_num=Erase_num+1
        ws0.cell(row =Erase_num +5, column = 2).value=float(line[6])
        ws0.cell(row =Erase_num +5, column = 4).value=float(line[10])
        
def FBC(lines):
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    global FBC_event
    FBC_event=FBC_event+1
    ws3.cell(row = FBC_event+5, column = 2).value=float(line[2])

def Program(lines):
    global PL,PU
    page=int(lines[10:12] ,16)
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    #print 'program'
    if(line[7]!='0'):
        
        if((page==0) or ((page % 2 !=0) and (page!=255))):
            PL=PL+1
            ws1.cell(row = PL+5, column = 2).value=float(line[7])
            ws1.cell(row = PL+5, column = 4).value=float(line[9])
                
        elif((page==255) or ((page % 2 ==0) and (page!=0))):
            PU=PU+1 
            ws1.cell(row = PU+5, column = 6).value=float(line[7])
            ws1.cell(row = PU+5, column = 8).value=float(line[9])
        
def Read(lines):
    global Read_num
    line= lines.replace('=',' ').replace('0x',' ').replace('h',' ').replace('(',' ').split()
    page=int(line[1][-2:],16)
    global RL, RU
    if((page==0) or ((page % 2 !=0) and (page!=255))):
        RL=RL+1
        ws2.cell(row = RL+5, column = 2).value=float(line[4])     
        
    elif((page==255) or ((page % 2 ==0) and (page!=0))):
        RU=RU+1 
        ws2.cell(row = RU+5, column = 4).value=float(line[4])

def init():
   
    ws0.cell('A2').value='Max'
    ws0.cell('A3').value='Min'
    ws0.cell('A4').value='Average'    
    ws0.cell('B1').value='Erase Loop'
    ws0.cell('D1').value='Erase Time'

    ws1.cell('A2').value='Max'
    ws1.cell('A3').value='Min'
    ws1.cell('A4').value='Average'
    ws1.cell('B1').value='P_L Loop'
    ws1.cell('D1').value='P_L Time'    
    ws1.cell('F1').value='P_U Loop'
    ws1.cell('H1').value='P_U Time' 
    
    ws2.cell('A2').value='Max'
    ws2.cell('A3').value='Min'
    ws2.cell('A4').value='Average'    
    ws2.cell('B1').value='RL Time'
    ws2.cell('D1').value='RU Time'
    
    ws3.cell('A2').value='Max'
    ws3.cell('A3').value='Min'
    ws3.cell('A4').value='Average'     
    ws3.cell('B1').value='FBC Check'
    ws3.cell('A2').value='FBC Max'
    ws3.cell('A3').value='FBC Min'
    ws3.cell('A4').value='FBC Average'
       
    ws.cell('C3').value='Max'
    ws.cell('D3').value='Min'
    ws.cell('E3').value='Average'
    ws.cell('B4').value='Erase Loop'
    ws.cell('B5').value='Erase Time'  
    ws.cell('B6').value='PL Loop'
    ws.cell('B7').value='PL Time'
    ws.cell('B8').value='PU Loop'
    ws.cell('B9').value='PU Time'   
    ws.cell('B10').value='Read Low'
    ws.cell('B11').value='Read Up'  
    ws.cell('B12').value='FBC'

def Cal():
    global FBC_event,PL,PU,RL,RU,Erase_num
    
    ws0.cell('B2').value="=max(B6:B"+str(5+Erase_num)+')'
    ws0.cell('B3').value="=min(B6:B"+str(5+Erase_num)+')'
    ws0.cell('B4').value="=average(B6:B"+str(5+Erase_num)+')'
    ws0.cell('D2').value="=max(D6:D"+str(5+Erase_num)+')'
    ws0.cell('D3').value="=min(D6:D"+str(5+Erase_num)+')'
    ws0.cell('D4').value="=average(D6:D"+str(5+Erase_num)+')'
    
    ws1.cell('B2').value="=max(B6:B"+str(5+PL)+')'
    ws1.cell('B3').value="=min(B6:B"+str(5+PL)+')'
    ws1.cell('B4').value="=average(B6:B"+str(5+PL)+')'
    ws1.cell('D2').value="=max(D6:D"+str(5+PL)+')'
    ws1.cell('D3').value="=min(D6:D"+str(5+PL)+')'
    ws1.cell('D4').value="=average(D6:D"+str(5+PL)+')'    
    ws1.cell('F2').value="=max(F6:F"+str(5+PU)+')'
    ws1.cell('F3').value="=min(F6:F"+str(5+PU)+')'
    ws1.cell('F4').value="=average(F6:F"+str(5+PU)+')'
    ws1.cell('H2').value="=max(H6:H"+str(5+PU)+')'
    ws1.cell('H3').value="=min(H6:H"+str(5+PU)+')'
    ws1.cell('H4').value="=average(H6:H"+str(5+PU)+')'              
    
    ws2.cell('B2').value="=max(B6:B"+str(5+RL)+')'
    ws2.cell('B3').value="=min(B6:B"+str(5+RL)+')'
    ws2.cell('B4').value="=average(B6:B"+str(5+RL)+')'
    ws2.cell('D2').value="=max(D6:D"+str(5+RU)+')'
    ws2.cell('D3').value="=min(D6:D"+str(5+RU)+')'
    ws2.cell('D4').value="=average(D6:D"+str(5+RU)+')'       
    
    ws3.cell('B2').value='=max'+"(B6:B"+str(5+FBC_event)+')'
    ws3.cell('B3').value='=min'+"(B6:B"+str(5+FBC_event)+')'
    ws3.cell('B4').value='=average'+"(B6:B"+str(5+FBC_event)+')'
    
         
def Summarize():
    global excel_name
    from xlrd import open_workbook,XL_CELL_TEXT
    book = open_workbook(excel_name)
    sheet0 = book.sheet_by_index(0)
    sheet1 = book.sheet_by_index(1)
      
    print sheet1.cell(0,1).value, sheet1.cell(1,0).value

    print sheet1.cell(0,0).value
    
def Summarize2():
    global excel_name
    from xlrd import open_workbook,XL_CELL_TEXT
    book = open_workbook(excel_name)
    sheet0 = book.sheet_by_index(0)
    sheet1 = book.sheet_by_index(1)
      
   
             
def main():  
    
    for i in range(1,10):
        file_iteration='UNIT'+str(i)+'-1.txt'    
                
        f = open(file_iteration, 'r')    #open T.txx
        for line in f:
            if(line[32:44]=='Program Loop'):
                Program(line)
                continue
                
            if(line[5:10]=='Total'):
                FBC(line)
                continue
                
            if ((line[2:7]=='Block') and (len(line)>20)):                
                Erase(line)
                continue
                
            if ((line[0:4]=='Page') and (len(line)<40)) :
                Read(line)
                continue
       
                             
            
if __name__ == '__main__':
    RL=0
    RU=0
    Erase_num=0
    FBC_event=0;
    Row_offset=4
    PL=0
    PU=0
    
#------------------Excel WorkBook Initiation----------------------------------    
    wb=Workbook()
    ws = wb.active
    ws.title="Summary"
    ws0 = wb.create_sheet()
    ws0.title="erase"
    ws1 = wb.create_sheet()
    ws1.title="program"
    ws2 = wb.create_sheet()
    ws2.title="read"
    ws3 = wb.create_sheet()
    ws3.title="FBC"
    excel_name='T data EFR_WEEK24&25.xlsx'
    init()
    main()
    Cal()
    
    wb.save(excel_name)
    Summarize()
    
