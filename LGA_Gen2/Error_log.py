__author__ = '20093'
'''
--IPython Code for Automatic Error log sorting------------

--By Jinqiang He,   March 11st, 2015

'''

from functools import partial
import glob
import os
from openpyxl import Workbook
import wx

# Get_path is refered from this site:
# http://stackoverflow.com/questions/9319317/quick-and-easy-file-dialog-in-python


def get_path(wildcard):
    app = wx.App(None)
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    dialog = wx.FileDialog(None, 'Open', wildcard=wildcard, style=style)
    if dialog.ShowModal() == wx.ID_OK:
        path = dialog.GetPath()
    else:
        path = None
    dialog.Destroy()
    return path

#------------------Excel WorkBook Initiation----------------------------------


def init():
    global ws1_row, ws1_column
    ws0.merge_cells('A2:D2')
    ws0.cell('A2').value = 'Error Log Stamp'
    ws0.merge_cells('E2:F2')
    ws0.cell('E2').value = 'Error Number(Hex)'
    ws0.merge_cells('G2:H2')
    ws0.cell('G2').value = 'Page Status'
    ws0.merge_cells('I2:L2')
    ws0.cell('I2').value = 'Cycle'
    ws0.cell('M2').value = 'Error Code'
    ws0.cell('N2').value = 'FIM'
    ws0.cell('O2').value = 'CE'
    ws0.cell('P2').value = 'Die'
    ws0.merge_cells('Q2:R2')
    ws0.cell('Q2').value = 'Block'
    ws0.merge_cells('S2:T2')
    ws0.cell('S2').value = 'Failing Page'

    ws0.merge_cells('Y2:AB2')
    ws0.cell('Y2').value = 'Read Error Stamp'
    ws0.merge_cells('AC2:AD2')
    ws0.cell('AC2').value = 'sector 1'
    ws0.merge_cells('AE2:AF2')
    ws0.cell('AE2').value = 'sector 2'
    ws0.merge_cells('AG2:AH2')
    ws0.cell('AG2').value = 'sector 3'
    ws0.merge_cells('AI2:AJ2')
    ws0.cell('AI2').value = 'sector 4'
    ws0.merge_cells('AK2:AL2')
    ws0.cell('AK2').value = 'sector 5'
    ws0.merge_cells('AM2:AN2')
    ws0.cell('AM2').value = 'sector 6'
    ws0.merge_cells('AO2:AP2')
    ws0.cell('AO2').value = 'sector 7'
    ws0.merge_cells('AQ2:AR2')
    ws0.cell('AQ2').value = 'sector 8'

    ws1.cell(row=ws1_row, column=ws1_col).value = "Error No."
    ws1.cell(row=ws1_row, column=ws1_col + 1).value = "Cycle"
    ws1.cell(row=ws1_row, column=ws1_col + 2).value = "Failure_Type"
    ws1.cell(row=ws1_row, column=ws1_col + 3).value = "status_type"
    ws1.cell(row=ws1_row, column=ws1_col + 4).value = "Channel/FIM"
    ws1.cell(row=ws1_row, column=ws1_col + 5).value = "Physical blk"
    ws1.cell(row=ws1_row, column=ws1_col + 6).value = "Page Number"
    ws1.cell(row=ws1_row, column=ws1_col + 7).value = "Page Type"

    ws1.cell(row=ws1_row, column=ws1_col + 8).value = 'sector 0'
    ws1.cell(row=ws1_row, column=ws1_col + 9).value = 'sector 1'
    ws1.cell(row=ws1_row, column=ws1_col + 10).value = 'sector 2'
    ws1.cell(row=ws1_row, column=ws1_col + 11).value = 'sector 3'
    ws1.cell(row=ws1_row, column=ws1_col + 12).value = 'sector 4'
    ws1.cell(row=ws1_row, column=ws1_col + 13).value = 'sector 5'
    ws1.cell(row=ws1_row, column=ws1_col + 14).value = 'sector 6'
    ws1.cell(row=ws1_row, column=ws1_col + 15).value = 'sector 7'


def main():
    errs = "Plane_Failure NO_LOG	NO_LOG_ERASE	MACRO_ERASE	MACRO_PROGRAM	MACRO_READ	LOWER_PAGE_PROGRAM_FAIL	UPPER_PAGE_PROGRAM_FAIL_WITH_LOWER_CORRUPTED	UPPER_PAGE_PROGRAM_FAIL	UPPER_PAGE_PROGRAM_FAIL_WITH_OTHERWL_CORRUPTED	DYN_RD_SUCCESS_OTHERWL_CORRUPTED	LOWER_PAGE_PROGRAM_FAIL_WITH_OTHERWL_CORRUPTED	SLC_PROGRAM_FAIL SLC_DYN_RD_FAIL	DYN_RD_CASE_WITH_DLA_ON_FAIL	DYN_RD_FAIL	GrownBB"

    labels = '0xD1 0xEF	0xEE	0x1	0x2	0x3	0x4	0x5	0x6	0x7	0x8	0x9	0xA	0xB	0xC	0xD	0xE'
    err_seq = errs.split()
    label_seq = [int(x, 16) for x in labels.split()]
    Errs_dict = {}
    global excel_name, file, Max_Cycle
    str = []
    bytes_per_line = 20
    line_num = 1
    err_num = 0
    Tech = raw_input(
        "Please input the plane spec(type 2 for 2 plane or 4 for 4 plane:")
    if (abs(int(Tech, 10) - 3) <> 1):
        print "Current Version only supports Ex2 2P or 4P, please restart the script"
        return 0
    die_stack = raw_input("Please input number of dies per chip: ")
    die_stack_int = int(die_stack, 10)
    if ((die_stack_int - 2) * (die_stack_int - 4) * (die_stack_int - 8) * (die_stack_int - 16) <> 0):
        print "Current Version only supports 2D,4D,8D,or 16D, please restart the script"
        return 0

    FIM_PerChip = 0
    CE_PerFim = 0
    Die_PerCE = 0
    blks_Per_Die = 0x0

    if (Tech == '2'):
        blks_Per_Die = 0x1000
        if(die_stack == '2'):
            FIM_PerChip = 2
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '4'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '8'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '16'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1

    elif(Tech == '4'):
        blks_Per_Die = 0x2000

        if(die_stack == '2'):
            FIM_PerChip = 2
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '4'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '8'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1
        elif(die_stack == '16'):
            FIM_PerChip = 4
            CE_PerFim = 1
            Die_PerCE = 1

    print "Please select the file for the error-log data to be processed."
    filename = get_path('*.pat')

    print "Processing File", filename
    excel_name = os.path.dirname(
        filename) + '/' + os.path.basename(filename)[:-4] + '--error log.xlsx'
    with open(filename, 'rb') as binary_file:
        for block in iter(partial(binary_file.read, bytes_per_line), ''):
            s1 = ' '.join('{0:02x}'.format(ord(b)) for b in block)
            line_num = line_num + 1
            ws.cell(row=line_num, column=1).value = s1
            if (s1[0:11] == '45 72 4c 67'):
                err_num = int(s1[4 * 3:4 * 3 + 2], 16) + \
                    int(s1[5 * 3:5 * 3 + 2], 16) * 0x100
                status_type = s1[6 * 3:6 * 3 + 2]
                column_offset = 1
                cycle = int(s1[11 * 3:11 * 3 + 2], 16) * 0x1000000 + int(s1[10 * 3:10 * 3 + 2], 16) * \
                    0x10000 + int(s1[9 * 3:9 * 3 + 2], 16) * \
                    0x100 + int(s1[8 * 3:8 * 3 + 2], 16)
                Error_Code_Num = int(s1[12 * 3:12 * 3 + 2], 16)
                Error_code = err_seq[label_seq.index(Error_Code_Num)]
                FIM = int(s1[13 * 3:13 * 3 + 2])
                CE = int(s1[14 * 3:14 * 3 + 2])
                Die = int(s1[15 * 3:15 * 3 + 2])
                blk = int(s1[16 * 3:16 * 3 + 2], 16) + \
                    int(s1[17 * 3:17 * 3 + 2], 16) * 0x100
                Page = int((s1[18 * 3:18 * 3 + 2]), 16) + \
                    int(s1[19 * 3:19 * 3 + 2], 16) * 0x100
                Phy_Die = CE * Die_PerCE + Die

                ws1.cell(row=err_num + 2, column=column_offset).value = err_num
                ws1.cell(
                    row=err_num + 2, column=column_offset + 1).value = cycle
                ws1.cell(
                    row=err_num + 2, column=column_offset + 2).value = Error_code
                ws1.cell(
                    row=err_num + 2, column=column_offset + 3).value = status_type
                ws1.cell(row=err_num + 2, column=column_offset + 4).value = FIM
                if Error_Code_Num == 10 or Error_Code_Num == 11:
                    ws1.cell(row=err_num + 2, column=column_offset +
                             5).value = hex(Phy_Die * blks_Per_Die + blk)
                    ws1.cell(
                        row=err_num + 2, column=column_offset + 6).value = hex(Page)
                    ws1.cell(
                        row=err_num + 2, column=column_offset + 7).value = 'SLC'
                else:
                    ws1.cell(row=err_num + 2, column=column_offset +
                             5).value = hex(Phy_Die * blks_Per_Die + blk)
                    ws1.cell(
                        row=err_num + 2, column=column_offset + 6).value = hex(Page)
                    if Page == 0 or (Page % 2 == 1 and Page <> 255):
                        Page_Type = 'LP'
                    else:
                        Page_Type = 'UP'
                    ws1.cell(
                        row=err_num + 2, column=column_offset + 7).value = Page_Type

            elif(s1[0:11] == '52 64 45 72'):
                column_offset = 25
                ws1.cell(row=err_num + 2, column=8 +
                         1).value = int(s1[5 * 3:5 * 3 + 2] + s1[4 * 3:4 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=9 +
                         1).value = int(s1[7 * 3:7 * 3 + 2] + s1[6 * 3:6 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=10 +
                         1).value = int(s1[9 * 3:9 * 3 + 2] + s1[8 * 3:8 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=11 +
                         1).value = int(s1[11 * 3:11 * 3 + 2] + s1[10 * 3:10 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=12 +
                         1).value = int(s1[13 * 3:13 * 3 + 2] + s1[12 * 3:12 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=13 +
                         1).value = int(s1[15 * 3:15 * 3 + 2] + s1[14 * 3:14 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=14 +
                         1).value = int(s1[17 * 3:17 * 3 + 2] + s1[16 * 3:16 * 3 + 2], 16)
                ws1.cell(row=err_num + 2, column=15 +
                         1).value = int(s1[19 * 3:19 * 3 + 2] + s1[18 * 3:18 * 3 + 2], 16)

            for i in range(20):
                ws0.cell(
                    row=err_num + 2, column=column_offset + i).value = s1[3 * i:3 * i + 2]

            for item in s1:
                str.append(item)

if __name__ == '__main__':
    excel_name = ' '
    file = ' '
    ws1_row = 1
    ws1_col = 1
    print "This API is for auto processing error-log data within pat files generated from CNE, and processed data will be saved in an excel file."
    print "Current Version only supports Raw Nand LGA60, will add support for other products gradually"
    print "by Jinqiang May 4th, 2015"
    print '****************************************************'
    print ' '

#------------------Excel WorkBook Initiation----------------------------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw Data"
    ws0 = wb.create_sheet()
    ws0.title = "Error Log Raw"
    ws1 = wb.create_sheet()
    ws1.title = "Error Log Processed"
    init()
    main()
    wb.save(excel_name)
    print "The processed Error Log Summary is saved as", excel_name
